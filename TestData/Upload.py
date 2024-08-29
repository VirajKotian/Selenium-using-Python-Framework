import time

from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions
from selenium.webdriver.support.wait import WebDriverWait

import openpyxl

def update_excel_data(filePath, searchTerm, colName, new_value):
    book = openpyxl.load_workbook(filePath)
    sheet = book.active
    Dict = {}
    for i in range(1, sheet.max_row+1):
        if sheet.cell(row = 1, column = i).value == colName:    #to get the first row
            Dict["col"] = i


    for i in range(1, sheet.max_column+1):           #To get Column
        for j in range(1, sheet.max_row+1):
            if sheet.cell(row=i, column=j).value == searchTerm:
                Dict["row"] = i


    sheet.cell(row = Dict["row"], column = Dict["col"]).value = new_value
    book.save(file_path)


file_path = "C:/Users/sanjeev kotian/Downloads/download.xlsx"

fruit_name = "Apple"
newValue = "700"

driver = webdriver.Chrome()
driver.implicitly_wait(5)
driver.get("https://rahulshettyacademy.com/upload-download-test/")
driver.find_element(By.ID,"downloadButton").click()


#Edit the Excel with Updated Value
update_excel_data(file_path, fruit_name, "price",newValue)

#To upload a file there must be a type 'file' attribute
file_input = driver.find_element(By.CSS_SELECTOR,"input[type='file']")
file_input.send_keys(file_path)

wait = WebDriverWait(driver, 5)
toast_locator = (By.CSS_SELECTOR,".Toastify__toast-body div:nth-child(2)")
wait.until(expected_conditions.visibility_of_element_located(toast_locator))
print(driver.find_element(*toast_locator).text)
price_column = driver.find_element(By.XPATH,"//div[text()='Price']").get_attribute("data-column-id")
Actual_price = driver.find_element(By.XPATH,"//div[text()= '"+fruit_name+"']/parent::div/parent::div/div[@id='cell-"+price_column+"-undefined']").text
print(Actual_price)
assert Actual_price == newValue


time.sleep(2)