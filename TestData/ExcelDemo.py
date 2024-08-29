import openpyxl
book = openpyxl.load_workbook("C:\\Users\\sanjeev kotian\\Downloads\\PythonDemo.xlsx")
sheet = book.active
Dict = {}
cell = sheet.cell(row = 1, column = 2)
print(cell.value)
sheet.cell(row = 2, column = 2).value = "Viraj"

print(sheet.cell(row = 2, column = 2).value)

print(sheet.max_row)

print(sheet.max_column)

print(sheet['A2'].value)             #can also print on basis of A5 B1 D2 ...

print("*******************************************************************************************************")

for i in range(1, sheet.max_row+1):
    if sheet.cell(row = i, column = 1).value:                  #To get row

        for j in range(1, sheet.max_column+1):                 #To get Column
            print(sheet.cell(row = i, column = j).value)

print("*******************************************************************************************************")

for i in range(1, sheet.max_row+1):
    if sheet.cell(row = i, column = 1).value == "Testcase2":    #To get row only Testcase 2

        for j in range(1, sheet.max_column+1):                 #To get Column
            print(sheet.cell(row = i, column = j).value)

print("*******************************************************************************************************")

for i in range(1, sheet.max_row+1):
    if sheet.cell(row = i, column = 1).value == "Testcase2":    #To get row only Testcase 2 without printing Testcase2

        for j in range(2, sheet.max_column+1):                 #To get Column

            Dict[sheet.cell(row = 1, column = j).value] = sheet.cell(row = i, column = j).value
            # Dict["firstname"] = "Viraj"
print(Dict)









