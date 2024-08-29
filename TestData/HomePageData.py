import openpyxl


class HomePageData:

    test_HomePage_data = [
                          {"Firstname": "Viraj","email":"hello@gmail.com", "password":"12345", "DOB":"24/05/2001", "Message":"HelloHR"},
                          {"Firstname":"Viruu", "email":"Chech@gmail.com", "password":"54321", "DOB": "24/05/2024","Message":"TakeCare"}
                          ]

    @staticmethod
    def getTestdata(test_case_name):
        Dict = {}
        book = openpyxl.load_workbook("C:\\Users\\sanjeev kotian\\Downloads\\PythonDemo.xlsx")
        sheet = book.active
        for i in range(1, sheet.max_row + 1):
            if sheet.cell(row=i, column=1).value == test_case_name:  # To get row only Testcase 2 without printing Testcase2

                for j in range(2, sheet.max_column + 1):  # To get Column

                    Dict[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value

        return[Dict]



    # ("Viruu", "Chech@gmail.com", "54321")]) #we can use self.driver.refesh() method also






